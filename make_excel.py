"""
Создаёт 2 Excel файла:
  participants.xlsx  — ID, Профессия, Навыки (для участников)
  login_keys.xlsx    — ID, Пароль (для администратора)
Затем проверяет что оба файла совпадают по числу и составу ID.
"""
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'http://185.73.126.238:8001'
DEFAULT_PASSWORD = 'fire2026'


def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def header_style(cell, bg='2563EB'):
    cell.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_style(cell, center=False):
    cell.font = Font(size=10, name='Calibri')
    cell.alignment = Alignment(
        horizontal='center' if center else 'left', vertical='center')
    thin = Side(style='thin', color='E2E8F0')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_participants(token: str) -> list[dict]:
    """Получаем всех участников через /directory."""
    r = requests.get(f'{BASE}/directory',
                     headers={'Authorization': f'Bearer {token}'})
    r.raise_for_status()
    participants = []
    for prof_group in r.json():
        for m in prof_group['members']:
            participants.append({
                'pid': m['pid'],
                'profession': prof_group['profession'],
                'skills': m.get('skills') or '—',
            })
    participants.sort(key=lambda x: x['pid'])
    return participants


def make_participants_xlsx(participants: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Участники'
    ws.row_dimensions[1].height = 24

    headers = ['ID', 'Профессия', 'Навыки / Специализация',
               'Имя (заполнить)', 'Telegram username', 'Заметки']
    widths   = [10,   22,          36,                    22,   20,                8]

    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        header_style(cell)
        col_width(ws, c, w)

    alt_fill = PatternFill('solid', fgColor='EFF6FF')
    for row_i, p in enumerate(participants, start=2):
        ws.row_dimensions[row_i].height = 18
        values = [p['pid'], p['profession'], p['skills'], '', '', '']
        fill = alt_fill if row_i % 2 == 0 else None
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=c, value=v)
            cell_style(cell, center=(c in [1]))
            if fill:
                cell.fill = fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(participants)+1}'
    wb.save(path)
    print(f'Saved: {path}  ({len(participants)} строк)')


def make_login_xlsx(participants: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ключи входа'
    ws.row_dimensions[1].height = 24

    headers = ['ID участника', 'Пароль (по умолчанию)', 'Пароль изменён?', 'Дата выдачи']
    widths   = [16,             22,                       18,                14]

    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        header_style(cell, bg='1D4ED8')
        col_width(ws, c, w)

    alt_fill = PatternFill('solid', fgColor='EFF6FF')
    for row_i, p in enumerate(participants, start=2):
        ws.row_dimensions[row_i].height = 18
        values = [p['pid'], DEFAULT_PASSWORD, 'Нет', '03.03.2026']
        fill = alt_fill if row_i % 2 == 0 else None
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=c, value=v)
            cell_style(cell, center=True)
            if fill:
                cell.fill = fill

    ws.freeze_panes = 'A2'
    wb.save(path)
    print(f'Saved: {path}  ({len(participants)} строк)')


def verify_files(path1: str, path2: str):
    wb1 = openpyxl.load_workbook(path1)
    wb2 = openpyxl.load_workbook(path2)
    ws1, ws2 = wb1.active, wb2.active

    ids1 = {ws1.cell(r, 1).value for r in range(2, ws1.max_row + 1)
            if ws1.cell(r, 1).value}
    ids2 = {ws2.cell(r, 1).value for r in range(2, ws2.max_row + 1)
            if ws2.cell(r, 1).value}

    print(f'\n=== Проверка таблиц ===')
    print(f'participants.xlsx : {len(ids1)} ID')
    print(f'login_keys.xlsx   : {len(ids2)} ID')
    only_in_1 = ids1 - ids2
    only_in_2 = ids2 - ids1
    if not only_in_1 and not only_in_2:
        print('OK - Таблицы совпадают по составу ID')
    else:
        if only_in_1:
            print(f'❌ Только в participants: {sorted(only_in_1)}')
        if only_in_2:
            print(f'❌ Только в login_keys: {sorted(only_in_2)}')


if __name__ == '__main__':
    # Логинимся как администратор
    r = requests.post(f'{BASE}/auth/login',
                      json={'pid': 'P-001', 'password': DEFAULT_PASSWORD})
    r.raise_for_status()
    token = r.json()['access_token']

    participants = build_participants(token)

    p1 = r'C:\Users\info\PycharmProjects\Pythonclaudebash\FIRE35\participants.xlsx'
    p2 = r'C:\Users\info\PycharmProjects\Pythonclaudebash\FIRE35\login_keys.xlsx'

    make_participants_xlsx(participants, p1)
    make_login_xlsx(participants, p2)
    verify_files(p1, p2)
